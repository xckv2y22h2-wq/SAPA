#!/usr/bin/env python3
"""
SAPA Attack Comparison Script

This script runs SAPA attacks across multiple models, datasets, and epsilon values,
then generates a comprehensive comparison Excel file from the results.

Usage:
    python run_semantic_attack_comparison.py --config configs/comparison_config.yaml
    python run_semantic_attack_comparison.py --config configs/comparison_config.yaml --show_config
"""

import os
import json
import argparse
import subprocess
import yaml
from datetime import datetime
from itertools import product

# Try to import pandas and openpyxl for Excel generation
try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    print("Warning: pandas not installed. Excel generation will be limited.")

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. Excel formatting will be limited.")


# ============================================================================
# Available Models and Datasets
# ============================================================================

AVAILABLE_MODELS = {
    'CLIP': {
        'description': 'Standard CLIP model (baseline)',
        'arch': ['ViT-B/32', 'ViT-B/16', 'ViT-L/14', 'RN50', 'RN101'],
    },
    'FARE': {
        'description': 'Feature-level Adversarial Robustness Enhancement',
        'arch': ['ViT-B/32'],
    },
    'TeCoA': {
        'description': 'Text-guided Contrastive Adversarial training',
        'arch': ['ViT-B/32'],
    },
    'PMG': {
        'description': 'Prompt-driven Multimodal Guidance',
        'arch': ['ViT-B/32'],
    },
    'TRADES': {
        'description': 'TRadeoff-inspired Adversarial DEfense via Surrogate-loss',
        'arch': ['ViT-B/32'],
    },
    'AUDIENCE': {
        'description': 'Adversarial Unsupervised Domain-Invariant Encoder',
        'arch': ['ViT-B/32'],
    },
    'MobileCLIP': {
        'description': 'MobileCLIP efficient model',
        'arch': ['S0', 'S1', 'S2'],
    },
    'MobileCLIP2': {
        'description': 'MobileCLIP2 efficient model',
        'arch': ['S0', 'S1', 'S2'],
    },
}

AVAILABLE_DATASETS = {
    'cifar10': {
        'description': 'CIFAR-10 (10 classes, 32x32 images)',
        'num_classes': 10,
        'image_size': 32,
    },
    'cifar100': {
        'description': 'CIFAR-100 (100 classes, 32x32 images)',
        'num_classes': 100,
        'image_size': 32,
    },
    'STL10': {
        'description': 'STL-10 (10 classes, 96x96 images)',
        'num_classes': 10,
        'image_size': 96,
    },
    'ImageNet': {
        'description': 'ImageNet-100 (100 classes, 224x224 images)',
        'num_classes': 100,
        'image_size': 224,
    },
    'Caltech101': {
        'description': 'Caltech-101 (101 object categories)',
        'num_classes': 101,
        'image_size': 224,
    },
    'Caltech256': {
        'description': 'Caltech-256 (256 object categories)',
        'num_classes': 256,
        'image_size': 224,
    },
    'OxfordPets': {
        'description': 'Oxford-IIIT Pet Dataset (37 pet breeds)',
        'num_classes': 37,
        'image_size': 224,
    },
    'Flowers102': {
        'description': 'Oxford 102 Flower Dataset',
        'num_classes': 102,
        'image_size': 224,
    },
    'Food101': {
        'description': 'Food-101 Dataset (101 food categories)',
        'num_classes': 101,
        'image_size': 224,
    },
    'DTD': {
        'description': 'Describable Textures Dataset (47 texture categories)',
        'num_classes': 47,
        'image_size': 224,
    },
    'EuroSAT': {
        'description': 'EuroSAT Satellite Image Dataset (10 classes)',
        'num_classes': 10,
        'image_size': 224,
    },
    'StanfordCars': {
        'description': 'Stanford Cars Dataset (196 car models)',
        'num_classes': 196,
        'image_size': 224,
    },
    'SUN397': {
        'description': 'SUN397 Scene Dataset (397 scene categories)',
        'num_classes': 397,
        'image_size': 224,
    },
    'FGVCAircraft': {
        'description': 'FGVC Aircraft Dataset (100 aircraft variants)',
        'num_classes': 100,
        'image_size': 224,
    },
    'Country211': {
        'description': 'Country-211 Geolocation Dataset',
        'num_classes': 211,
        'image_size': 224,
    },
    'PCAM': {
        'description': 'PatchCamelyon Medical Dataset (2 classes)',
        'num_classes': 2,
        'image_size': 224,
    },
}


# ============================================================================
# Default Configuration
# ============================================================================

DEFAULT_CONFIG = {
    'models': ['CLIP', 'FARE', 'TeCoA'],
    'datasets': ['cifar10', 'cifar100', 'ImageNet'],
    'epsilons': [1, 2, 4, 8],
    'attack': {
        'num_test_samples': 100,
        'attack_iters': 30,
        'alpha': 0.01,
        'tta_frequency': 10,
        'batch_size': 32,
    },
    'model_settings': {
        'arch': 'ViT-B/32',
        'model_dir': './models/',
        'mobileclip_dir': './models/mobileclip_models/',
    },
    'data': {
        'data_path': './datasets/',
        'workers': 4,
    },
    'hardware': {
        'gpu': 0,
        'seed': 42,
    },
    'output': {
        'output_dir': './results',
        'save_images': False,
    },
}


# ============================================================================
# Configuration Functions
# ============================================================================

def create_default_config(output_path):
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    with open(output_path, 'w') as f:
        f.write("# SAPA Attack Comparison Configuration\n")
        f.write("# =====================================\n\n")
        f.write(f"# Available models: {list(AVAILABLE_MODELS.keys())}\n")
        f.write(f"# Available datasets: {list(AVAILABLE_DATASETS.keys())}\n\n")
        yaml.dump(DEFAULT_CONFIG, f, default_flow_style=False, sort_keys=False)
    
    print(f"Created default config at: {output_path}")
    return DEFAULT_CONFIG


def load_config(config_path):
    if not os.path.exists(config_path):
        print(f"Config file not found: {config_path}")
        print("Creating default configuration...")
        return create_default_config(config_path)
    
    with open(config_path, 'r') as f:
        config = yaml.safe_load(f)
    
    # Handle empty or invalid config file
    if config is None:
        print(f"Warning: Config file is empty or invalid: {config_path}")
        print("Using default configuration...")
        return DEFAULT_CONFIG.copy()
    
    # Merge with defaults for missing keys
    merged = DEFAULT_CONFIG.copy()
    
    def deep_update(base, update):
        if update is None:
            return
        for key, value in update.items():
            if key.startswith('_'):
                continue
            if isinstance(value, dict) and key in base and isinstance(base[key], dict):
                deep_update(base[key], value)
            else:
                base[key] = value
    
    deep_update(merged, config)
    
    return merged


def validate_config(config):
    errors = []
    warnings = []
    
    for model in config['models']:
        if model not in AVAILABLE_MODELS:
            errors.append(f"Unknown model: {model}")
    
    for dataset in config['datasets']:
        if dataset not in AVAILABLE_DATASETS:
            errors.append(f"Unknown dataset: {dataset}")
    
    for eps in config['epsilons']:
        if eps <= 0 or eps > 255:
            errors.append(f"Invalid epsilon: {eps}/255 (must be between 1 and 255)")
    
    data_path = config['data']['data_path']
    if not os.path.exists(data_path):
        warnings.append(f"Data path does not exist: {data_path}")
    
    model_dir = config['model_settings']['model_dir']
    if not os.path.exists(model_dir):
        warnings.append(f"Model directory does not exist: {model_dir}")
    
    if errors:
        print("\nConfiguration errors:")
        for e in errors:
            print(f"  - {e}")
    
    if warnings:
        print("\n Configuration warnings:")
        for w in warnings:
            print(f"  - {w}")
    
    return len(errors) == 0


def print_config(config):
    print("\n" + "="*70)
    print("CONFIGURATION SUMMARY")
    print("="*70)
    
    print(f"\nAvailable Models:")
    for name, info in AVAILABLE_MODELS.items():
        marker = "✓" if name in config['models'] else " "
        print(f"  [{marker}] {name}: {info['description']}")
    
    print(f"\nAvailable Datasets:")
    for name, info in AVAILABLE_DATASETS.items():
        marker = "✓" if name in config['datasets'] else " "
        print(f"  [{marker}] {name}: {info['description']}")
    
    print(f"\nSelected Models ({len(config['models'])}):")
    for model in config['models']:
        print(f"  - {model}")
    
    print(f"\nSelected Datasets ({len(config['datasets'])}):")
    for dataset in config['datasets']:
        print(f"  - {dataset}")
    
    print(f"\nEpsilon values ({len(config['epsilons'])}):")
    for eps in config['epsilons']:
        print(f"  - {eps}/255 = {eps/255:.6f}")
    
    print(f"\nAttack parameters:")
    for key, value in config['attack'].items():
        print(f"  - {key}: {value}")
    
    print(f"\nPaths:")
    print(f"  - data_path: {config['data']['data_path']}")
    print(f"  - model_dir: {config['model_settings']['model_dir']}")
    print(f"  - output_dir: {config['output']['output_dir']}")
    
    total_runs = len(config['models']) * len(config['datasets']) * len(config['epsilons'])
    print(f"\nTotal configurations to run: {total_runs}")
    print("="*70 + "\n")


# ============================================================================
# Attack Runner
# ============================================================================

def run_single_attack(model, dataset, epsilon, config):
    epsilon_float = epsilon / 255.0

    cmd = [
        'python', 'run_semantic_attack_config.py',
        '--model', model,
        '--dataset', dataset,
        '--epsilon', f"{epsilon_float:.6f}",
        '--num_test_samples', str(config['attack']['num_test_samples']),
        '--attack_iters', str(config['attack']['attack_iters']),
        '--alpha', str(config['attack']['alpha']),
        '--tta_frequency', str(config['attack']['tta_frequency']),
        '--batch_size', str(config['attack']['batch_size']),
        '--workers', str(config['data']['workers']),
        '--gpu', str(config['hardware']['gpu']),
        '--data_path', config['data']['data_path'],
        '--model_dir', config['model_settings']['model_dir'],
        '--output_dir', config['output']['output_dir'],
        '--arch', config['model_settings']['arch'],
    ]

    # Priority 1B: Add cross-model STA flag if enabled in config
    if config['output'].get('compute_cross_model_sta', False):
        cmd.append('--compute_cross_model_sta')
        eval_model = config['output'].get('eval_model', 'openclip_vit_l')
        cmd.extend(['--eval_model', eval_model])
    
    print(f"\n{'='*70}")
    print(f"Running: {model} on {dataset} with ε={epsilon}/255")
    print(f"{'='*70}")
    
    try:
        result = subprocess.run(cmd, capture_output=False, text=True)
        return result.returncode == 0
    except Exception as e:
        print(f"Error running attack: {e}")
        return False


def run_all_attacks(config):
    models = config['models']
    datasets = config['datasets']
    epsilons = config['epsilons']
    
    total = len(models) * len(datasets) * len(epsilons)
    completed = 0
    failed = []
    
    print(f"\n{'='*70}")
    print(f"SAPA ATTACK COMPARISON - RUNNING ATTACKS")
    print(f"{'='*70}")
    print(f"Models: {models}")
    print(f"Datasets: {datasets}")
    print(f"Epsilons: {[f'{e}/255' for e in epsilons]}")
    print(f"Total configurations: {total}")
    print(f"{'='*70}\n")
    
    for model, dataset, epsilon in product(models, datasets, epsilons):
        success = run_single_attack(model, dataset, epsilon, config)
        completed += 1
        
        if not success:
            failed.append((model, dataset, epsilon))
        
        print(f"\nProgress: {completed}/{total} ({completed/total*100:.1f}%)")
    
    print(f"\n{'='*70}")
    print(f"ATTACK RUNS COMPLETE")
    print(f"{'='*70}")
    print(f"Total: {total}")
    print(f"Successful: {total - len(failed)}")
    print(f"Failed: {len(failed)}")
    
    if failed:
        print(f"\nFailed configurations:")
        for m, d, e in failed:
            print(f"  - {m} / {d} / ε={e}/255")
    
    return {
        'total': total,
        'successful': total - len(failed),
        'failed': failed
    }


# ============================================================================
# Result Collection
# ============================================================================

def collect_results(config):
    output_dir = config['output']['output_dir']
    models = config['models']
    datasets = config['datasets']
    epsilons = config['epsilons']
    
    results = []
    missing = []
    
    for model, dataset, epsilon in product(models, datasets, epsilons):
        epsilon_float = epsilon / 255.0
        
        filename = f'sapa_{model}_{dataset}_eps{epsilon_float:.4f}_results.json'
        filepath = os.path.join(output_dir, filename)
        
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r') as f:
                    data = json.load(f)
                
                summary = data.get('summary', {})
                
                result = {
                    'model': model,
                    'dataset': dataset,
                    'epsilon': epsilon,
                    'epsilon_float': epsilon_float,
                    'epsilon_str': f'{epsilon}/255',
                    'success_rate': summary.get('success_rate', 0),
                    'avg_semantic_alignment': summary.get('avg_semantic_alignment', 0),
                    'avg_l_inf_norm': summary.get('avg_l_inf_norm', 0),
                    'avg_l2_norm': summary.get('avg_l2_norm', 0),
                    'total_samples': summary.get('total_samples', 0),
                    'successfully_attacked': summary.get('successfully_attacked', 0),
                    'skipped_no_anchor': summary.get('skipped_no_anchor', 0),
                    'skipped_errors': summary.get('skipped_errors', 0),
                }
                results.append(result)
                
            except Exception as e:
                print(f"Error reading {filepath}: {e}")
                missing.append((model, dataset, epsilon))
        else:
            missing.append((model, dataset, epsilon))
    
    print(f"\nCollected {len(results)} result files")
    if missing:
        print(f"Missing {len(missing)} result files:")
        for m, d, e in missing[:10]:
            print(f"  - {m} / {d} / ε={e}/255")
        if len(missing) > 10:
            print(f"  ... and {len(missing) - 10} more")
    
    return results


# ============================================================================
# Excel Generation
# ============================================================================

def generate_excel_report(results, output_path, config):
    if not HAS_PANDAS:
        print("Error: pandas is required for Excel generation")
        print("Install with: pip install pandas openpyxl")
        generate_csv_report(results, os.path.dirname(output_path))
        return
    
    if not results:
        print("No results to generate report from")
        return
    
    df = pd.DataFrame(results)
    eps_order = [f'{e}/255' for e in sorted(config['epsilons'])]
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        
        # Sheet 1: Raw Data
        df.to_excel(writer, sheet_name='Raw Data', index=False)
        
        # Sheet 2: Success Rate by Model x Dataset (for each epsilon)
        for epsilon in config['epsilons']:
            eps_str = f'{epsilon}_255'
            sheet_name = f'Success_eps_{eps_str}'
            
            df_eps = df[df['epsilon'] == epsilon]
            if not df_eps.empty:
                pivot = df_eps.pivot_table(
                    values='success_rate',
                    index='model',
                    columns='dataset',
                    aggfunc='first'
                )
                pivot.to_excel(writer, sheet_name=sheet_name)
        
        # Sheet 3: Success Rate Summary (Model x Epsilon)
        summary_model_eps = df.groupby(['model', 'epsilon_str']).agg({
            'success_rate': 'mean',
            'avg_semantic_alignment': 'mean',
        }).reset_index()
        
        pivot_model_eps = summary_model_eps.pivot_table(
            values='success_rate',
            index='model',
            columns='epsilon_str',
            aggfunc='first'
        )
        pivot_model_eps = pivot_model_eps.reindex(
            columns=[c for c in eps_order if c in pivot_model_eps.columns]
        )
        pivot_model_eps.to_excel(writer, sheet_name='Success_Model_Epsilon')
        
        # Sheet 4: Success Rate Summary (Dataset x Epsilon)
        summary_dataset_eps = df.groupby(['dataset', 'epsilon_str']).agg({
            'success_rate': 'mean',
            'avg_semantic_alignment': 'mean',
        }).reset_index()
        
        pivot_dataset_eps = summary_dataset_eps.pivot_table(
            values='success_rate',
            index='dataset',
            columns='epsilon_str',
            aggfunc='first'
        )
        pivot_dataset_eps = pivot_dataset_eps.reindex(
            columns=[c for c in eps_order if c in pivot_dataset_eps.columns]
        )
        pivot_dataset_eps.to_excel(writer, sheet_name='Success_Dataset_Epsilon')
        
        # Sheet 5: Semantic Alignment by Model x Dataset
        pivot_alignment = df.pivot_table(
            values='avg_semantic_alignment',
            index='model',
            columns='dataset',
            aggfunc='mean'
        )
        pivot_alignment.to_excel(writer, sheet_name='Semantic_Alignment')
        
        # Sheet 6: Full Comparison Table
        comparison_df = df.pivot_table(
            values=['success_rate', 'avg_semantic_alignment', 'avg_l_inf_norm'],
            index=['model', 'dataset'],
            columns='epsilon_str',
            aggfunc='first'
        )
        comparison_df.to_excel(writer, sheet_name='Full_Comparison')
        
        # Sheet 7: Statistics Summary
        stats_data = []
        for model in df['model'].unique():
            for dataset in df['dataset'].unique():
                subset = df[(df['model'] == model) & (df['dataset'] == dataset)]
                if not subset.empty:
                    stats_data.append({
                        'Model': model,
                        'Dataset': dataset,
                        'Avg Success Rate': subset['success_rate'].mean(),
                        'Max Success Rate': subset['success_rate'].max(),
                        'Min Success Rate': subset['success_rate'].min(),
                        'Avg Semantic Align': subset['avg_semantic_alignment'].mean(),
                        'Total Samples': subset['total_samples'].sum(),
                        'Total Attacked': subset['successfully_attacked'].sum(),
                    })
        
        stats_df = pd.DataFrame(stats_data)
        stats_df.to_excel(writer, sheet_name='Statistics', index=False)
        
        # Sheet 8: Configuration
        config_data = [
            {'Parameter': 'Models', 'Value': ', '.join(config['models'])},
            {'Parameter': 'Datasets', 'Value': ', '.join(config['datasets'])},
            {'Parameter': 'Epsilons', 'Value': ', '.join([f'{e}/255' for e in config['epsilons']])},
            {'Parameter': 'Test Samples', 'Value': config['attack']['num_test_samples']},
            {'Parameter': 'Attack Iterations', 'Value': config['attack']['attack_iters']},
            {'Parameter': 'Alpha', 'Value': config['attack']['alpha']},
            {'Parameter': 'TTA Frequency', 'Value': config['attack']['tta_frequency']},
            {'Parameter': 'Architecture', 'Value': config['model_settings']['arch']},
            {'Parameter': 'Generated', 'Value': datetime.now().strftime('%Y-%m-%d %H:%M:%S')},
        ]
        config_df = pd.DataFrame(config_data)
        config_df.to_excel(writer, sheet_name='Configuration', index=False)
    
    print(f"\n Excel report saved to: {output_path}")
    
    if HAS_OPENPYXL:
        format_excel_report(output_path)


def format_excel_report(filepath):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.cell import MergedCell
    
    wb = load_workbook(filepath)
    
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format header row (skip merged cells)
        for cell in ws[1]:
            if not isinstance(cell, MergedCell):
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
                cell.border = border
        
        # Format data rows (skip merged cells)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center')
                    
                    if isinstance(cell.value, float) and 0 <= cell.value <= 1:
                        cell.number_format = '0.00%'
        
        # Adjust column widths using column index instead of column_letter
        for col_idx, column in enumerate(ws.columns, start=1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            
            for cell in column:
                # Skip merged cells for width calculation
                if isinstance(cell, MergedCell):
                    continue
                try:
                    if cell.value is not None and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 30)
            if adjusted_width > 0:
                ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(filepath)
    print(f" Applied formatting to Excel report")


def generate_csv_report(results, output_dir):
    import csv
    
    raw_path = os.path.join(output_dir, 'sapa_comparison_raw.csv')
    with open(raw_path, 'w', newline='') as f:
        if results:
            writer = csv.DictWriter(f, fieldnames=results[0].keys())
            writer.writeheader()
            writer.writerows(results)
    print(f" Raw data saved to: {raw_path}")
    
    summary_path = os.path.join(output_dir, 'sapa_comparison_summary.csv')
    with open(summary_path, 'w', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(['Model', 'Dataset', 'Epsilon', 'Success Rate', 'Semantic Alignment'])
        for r in results:
            writer.writerow([
                r['model'],
                r['dataset'],
                r['epsilon_str'],
                f"{r['success_rate']:.2%}",
                f"{r['avg_semantic_alignment']:.4f}"
            ])
    print(f"✓ Summary saved to: {summary_path}")


# ============================================================================
# Main
# ============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description='SAPA Attack Comparison',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Run attacks and generate Excel
  python run_semantic_attack_comparison.py --config configs/comparison_config.yaml

  # Show configuration only
  python run_semantic_attack_comparison.py --config configs/comparison_config.yaml --show_config
        """
    )
    
    parser.add_argument('--config', type=str, required=True,
                        help='Path to YAML configuration file')
    parser.add_argument('--show_config', action='store_true',
                        help='Show configuration and exit (do not run attacks)')
    
    return parser.parse_args()


def main():
    args = parse_args()
    
    # Load configuration
    config = load_config(args.config)
    
    # Validate configuration
    if not validate_config(config):
        print("\nConfiguration validation failed. Please fix errors and try again.")
        return
    
    # Print configuration summary
    print_config(config)
    
    # Show config only mode
    if args.show_config:
        return
    
    # Ensure output directory exists
    os.makedirs(config['output']['output_dir'], exist_ok=True)
    
    # Step 1: Run all attacks
    print("\n" + "="*70)
    print("STEP 1: RUNNING ATTACKS")
    print("="*70)
    run_all_attacks(config)
    
    # Step 2: Collect results and generate Excel
    print("\n" + "="*70)
    print("STEP 2: GENERATING EXCEL REPORT")
    print("="*70)
    results = collect_results(config)
    
    if results:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_path = os.path.join(
            config['output']['output_dir'],
            f'sapa_comparison_{timestamp}.xlsx'
        )
        generate_excel_report(results, excel_path, config)
    else:
        print("No results found. Excel report not generated.")
    
    print("\n" + "="*70)
    print("COMPARISON COMPLETE")
    print("="*70)


if __name__ == "__main__":
    main()
