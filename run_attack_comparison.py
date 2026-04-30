#!/usr/bin/env python3
"""
Attack Comparison Script

This script runs various adversarial attacks across multiple models, datasets, and epsilon values,
then generates a comprehensive comparison Excel file from the results.

Usage:
    python run_attack_comparison.py --config config.yaml
    python run_attack_comparison.py --config config.yaml --show_config
"""

import os
import subprocess
import argparse
import time
import json
import yaml
import pandas as pd
from pathlib import Path
from itertools import product

# ============================================================================
# Available Models and Datasets
# ============================================================================

AVAILABLE_MODELS = {
    'CLIP': {
        'description': 'Standard CLIP model (baseline)',
        'arch': ['ViT-B/32', 'ViT-B/16', 'ViT-L/14', 'RN50', 'RN101'],
    },
    'FARE': {
        'description': 'Feature-level Adversarial Robustness Enhancement',
        'arch': ['ViT-B/32'],
    },
    'TeCoA': {
        'description': 'Text-guided Contrastive Adversarial training',
        'arch': ['ViT-B/32'],
    },
    'PMG': {
        'description': 'Prompt-driven Multimodal Guidance',
        'arch': ['ViT-B/32'],
    },
    'TRADES': {
        'description': 'TRadeoff-inspired Adversarial DEfense via Surrogate-loss',
        'arch': ['ViT-B/32'],
    },
    'AUDIENCE': {
        'description': 'Adversarial Unsupervised Domain-Invariant Encoder',
        'arch': ['ViT-B/32'],
    },
    'MobileCLIP': {
        'description': 'MobileCLIP efficient model',
        'arch': ['S0', 'S1', 'S2'],
    },
    'MobileCLIP2': {
        'description': 'MobileCLIP2 efficient model',
        'arch': ['S0', 'S1', 'S2'],
    },
}

AVAILABLE_DATASETS = {
    'cifar10': {'description': 'CIFAR-10 (10 classes)', 'num_classes': 10},
    'cifar100': {'description': 'CIFAR-100 (100 classes)', 'num_classes': 100},
    'STL10': {'description': 'STL-10 (10 classes)', 'num_classes': 10},
    'ImageNet': {'description': 'ImageNet-100 (100 classes)', 'num_classes': 100},
    'Caltech101': {'description': 'Caltech-101 (101 categories)', 'num_classes': 101},
    'Caltech256': {'description': 'Caltech-256 (256 categories)', 'num_classes': 256},
    'OxfordPets': {'description': 'Oxford-IIIT Pet (37 breeds)', 'num_classes': 37},
    'Flowers102': {'description': 'Oxford 102 Flower', 'num_classes': 102},
    'Food101': {'description': 'Food-101 (101 categories)', 'num_classes': 101},
    'DTD': {'description': 'Describable Textures (47 categories)', 'num_classes': 47},
    'EuroSAT': {'description': 'EuroSAT Satellite (10 classes)', 'num_classes': 10},
    'StanfordCars': {'description': 'Stanford Cars (196 models)', 'num_classes': 196},
    'SUN397': {'description': 'SUN397 Scene (397 categories)', 'num_classes': 397},
    'FGVCAircraft': {'description': 'FGVC Aircraft (100 variants)', 'num_classes': 100},
    'Country211': {'description': 'Country-211 Geolocation', 'num_classes': 211},
    'PCAM': {'description': 'PatchCamelyon Medical (2 classes)', 'num_classes': 2},
}

AVAILABLE_ATTACKS = ['pgd', 'CW', 'targeted_pgd', 'targeted_cw', 'sapa']

# ============================================================================
# Default Configuration
# ============================================================================

DEFAULT_CONFIG = {
    'models': ['CLIP', 'FARE', 'TeCoA'],
    'datasets': ['cifar10', 'ImageNet'],
    'attacks': ['pgd', 'CW', 'sapa'],
    'epsilons': [1, 2, 4, 8],  # In 1/255 units
    'attack': {
        'attack_iters': 20,

        'embedding_loss_weight': 0.5,
        'semantic_anchor_weight': 0.3,
        'contrastive_loss_weight': 0.2,
        'adversarial_loss_weight': 0.1,
    },
    'model_settings': {
        'arch': 'ViT-B/32',
        'model_dir': './models/',
    },
    'data': {
        'data_path': './datasets/',
    },
    'output': {
        'output_dir': './attack_comparison_results',
        'visualize': False,
    },
}


# ============================================================================
# Configuration Functions
# ============================================================================

def create_default_config(output_path):
    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else '.', exist_ok=True)
    
    with open(output_path, 'w') as f:
        f.write("# Attack Comparison Configuration\n")
        f.write("# ================================\n\n")
        f.write(f"# Available models: {list(AVAILABLE_MODELS.keys())}\n")
        f.write(f"# Available datasets: {list(AVAILABLE_DATASETS.keys())}\n")
        f.write(f"# Available attacks: {AVAILABLE_ATTACKS}\n\n")
        yaml.dump(DEFAULT_CONFIG, f, default_flow_style=False, sort_keys=False)
    
    print(f"✓ Created default config at: {output_path}")
    return DEFAULT_CONFIG


def load_config(config_path):
    if not os.path.exists(config_path):
        print(f"Config file not found: {config_path}")
        print("Creating default configuration...")
        return create_default_config(config_path)
    
    with open(config_path, 'r') as f:
        config = yaml.safe_load(f)
    
    # Handle empty or invalid config file
    if config is None:
        print(f"Warning: Config file is empty or invalid: {config_path}")
        print("Using default configuration...")
        return DEFAULT_CONFIG.copy()
    
    # Merge with defaults for missing keys
    merged = DEFAULT_CONFIG.copy()
    
    def deep_update(base, update):
        if update is None:
            return
        for key, value in update.items():
            if key.startswith('_'):
                continue
            if isinstance(value, dict) and key in base and isinstance(base[key], dict):
                deep_update(base[key], value)
            else:
                base[key] = value
    
    deep_update(merged, config)
    
    return merged


def validate_config(config):
    errors = []
    warnings = []
    
    for model in config['models']:
        if model not in AVAILABLE_MODELS:
            errors.append(f"Unknown model: {model}")
    
    for dataset in config['datasets']:
        if dataset not in AVAILABLE_DATASETS:
            errors.append(f"Unknown dataset: {dataset}")
    
    for attack in config.get('attacks', []):
        if attack not in AVAILABLE_ATTACKS:
            warnings.append(f"Unknown attack: {attack}")
    
    for eps in config['epsilons']:
        if eps <= 0 or eps > 255:
            errors.append(f"Invalid epsilon: {eps}/255 (must be between 1 and 255)")
    
    data_path = config['data']['data_path']
    if not os.path.exists(data_path):
        warnings.append(f"Data path does not exist: {data_path}")
    
    model_dir = config['model_settings']['model_dir']
    if not os.path.exists(model_dir):
        warnings.append(f"Model directory does not exist: {model_dir}")
    
    if errors:
        print("\n❌ Configuration errors:")
        for e in errors:
            print(f"  - {e}")
    
    if warnings:
        print("\n⚠ Configuration warnings:")
        for w in warnings:
            print(f"  - {w}")
    
    return len(errors) == 0


def print_config(config):
    print("\n" + "="*70)
    print("CONFIGURATION SUMMARY")
    print("="*70)
    
    print(f"\nSelected Models ({len(config['models'])}):")
    for model in config['models']:
        info = AVAILABLE_MODELS.get(model, {})
        print(f"  - {model}: {info.get('description', 'N/A')}")
    
    print(f"\nSelected Datasets ({len(config['datasets'])}):")
    for dataset in config['datasets']:
        info = AVAILABLE_DATASETS.get(dataset, {})
        print(f"  - {dataset}: {info.get('description', 'N/A')}")
    
    print(f"\nSelected Attacks ({len(config.get('attacks', []))}):")
    for attack in config.get('attacks', []):
        print(f"  - {attack}")
    
    print(f"\nEpsilon values ({len(config['epsilons'])}):")
    for eps in config['epsilons']:
        print(f"  - {eps}/255 = {eps/255:.6f}")
    
    print(f"\nAttack parameters:")
    for key, value in config['attack'].items():
        print(f"  - {key}: {value}")
    
    print(f"\nPaths:")
    print(f"  - data_path: {config['data']['data_path']}")
    print(f"  - model_dir: {config['model_settings']['model_dir']}")
    print(f"  - output_dir: {config['output']['output_dir']}")
    
    total_runs = len(config['models']) * len(config['datasets']) * len(config.get('attacks', [])) * len(config['epsilons'])
    print(f"\nTotal configurations to run: {total_runs}")
    print("="*70 + "\n")


def parse_args():
    parser = argparse.ArgumentParser(description='Compare different attacks across multiple datasets')
  
    parser.add_argument('--config', type=str, default='config.yaml',
                        help='Path to YAML configuration file (default: config.yaml)')
    parser.add_argument('--show_config', action='store_true',
                        help='Show configuration summary and exit')
    parser.add_argument('--skip_existing', action='store_true', default=True,
                        help='Skip attacks if result files already exist (default: True)')
    parser.add_argument('--force', action='store_true',
                        help='Force re-run all attacks even if results exist')
    parser.add_argument('--force_clean', action='store_true',
                        help='Force re-evaluation of clean accuracy (ignores cache)')

    # Override options (take precedence over config file)
    parser.add_argument('--models', type=str, nargs='+',
                        help='Override models from config')
    parser.add_argument('--datasets', type=str, nargs='+',
                        help='Override datasets from config')
    parser.add_argument('--attacks', type=str, nargs='+',
                        help='Override attacks from config')
    parser.add_argument('--epsilons', type=int, nargs='+',
                        help='Override epsilons from config (in 1/255 units)')

    args = parser.parse_args()
    return args


def get_attack_specific_params(attack, config, epsilon_float=None):
    params = []
    attack_config = config.get('attack', {})

    # Compute stepsize from epsilon if not manually specified
    if 'stepsize' in attack_config:
        # Use manually specified stepsize (overrides auto-computation)
        stepsize = attack_config['stepsize']
    elif epsilon_float is not None:
        # Auto-compute: stepsize = 2.5 * epsilon / iterations (Madry et al. formula)
        attack_iters = attack_config.get('attack_iters', 20)
        stepsize_factor = attack_config.get('stepsize_factor', 2.5)  # Allow customization
        stepsize = stepsize_factor * epsilon_float / attack_iters
        print(f"  Auto-computed stepsize: {stepsize:.6f} (= {stepsize_factor} * {epsilon_float:.4f} / {attack_iters})")
    else:
        stepsize = None

    if stepsize is not None:
        params.extend(['--stepsize', str(stepsize)])

    # SAPA-specific parameters
    params.extend([
        '--embedding_loss_weight', str(attack_config.get('embedding_loss_weight', 0.5)),
        '--semantic_anchor_weight', str(attack_config.get('semantic_anchor_weight', 0.3)),
        '--contrastive_loss_weight', str(attack_config.get('contrastive_loss_weight', 0.2)),
        '--adversarial_loss_weight', str(attack_config.get('adversarial_loss_weight', 1.0)),
        '--semantic_alignment_weight', str(attack_config.get('semantic_alignment_weight', 2.0)),
        '--semantic_strategy', str(attack_config.get('semantic_strategy', 'similar')),
        '--tta_frequency', str(attack_config.get('tta_frequency', 10))
    ])

    # Add sapa_variant if specified
    if attack_config.get('sapa_variant'):
        params.extend(['--sapa_variant', attack_config.get('sapa_variant')])

    # Add adaptive_weights flag if specified
    if attack_config.get('adaptive_weights', False):
        params.append('--adaptive_weights')

    # Add layer_weights if specified
    if attack_config.get('layer_weights'):
        params.extend(['--layer_weights', attack_config.get('layer_weights')])

    return params


def get_attack_iterations(attack, config):
    return config.get('attack', {}).get('attack_iters', 20)


def get_available_models(model_dir):
    all_models = [
        {
            'type': 'CLIP',
            'path': model_dir,  # CLIP doesn't need checkpoint, but pass dir for consistency
            'feature_weight': 0.5,
            'description_weight': 0.3,
            'compositional_weight': 0.3,
            'adaptive_weight': 0.2,
            'image_weight': 0.6,
            'text_weight': 0.4
        },
        {
            'type': 'TeCoA',
            'path': model_dir,  # Pass directory, not full path
            'feature_weight': 0.2,
            'description_weight': 0.6,
            'compositional_weight': 0.4,
            'adaptive_weight': 0.3,
            'image_weight': 0.6,
            'text_weight': 0.4
        },
        {
            'type': 'PMG',
            'path': model_dir,  # Pass directory, not full path
            'feature_weight': 0.8,
            'description_weight': 0.1,
            'compositional_weight': 0.3,
            'adaptive_weight': 0.2,
            'image_weight': 0.6,
            'text_weight': 0.4
        },
        {
            'type': 'FARE',
            'path': model_dir,  # Pass directory, not full path
            'feature_weight': 0.8,
            'description_weight': 0.2,
            'compositional_weight': 0.3,
            'adaptive_weight': 0.2,
            'image_weight': 0.6,
            'text_weight': 0.4
        },
        {
            'type': 'AUDIENCE',
            'path': model_dir,  # Pass directory, not full path
            'feature_weight': 0.8,
            'description_weight': 0.2,
            'compositional_weight': 0.3,
            'adaptive_weight': 0.2,
            'image_weight': 0.6,
            'text_weight': 0.4
        },
        {
            'type': 'TRADES',
            'path': model_dir,  # Pass directory, not full path
            'feature_weight': 0.8,
            'description_weight': 0.2,
            'compositional_weight': 0.3,
            'adaptive_weight': 0.2,
            'image_weight': 0.6,
            'text_weight': 0.4
        },
        {
            'type': 'MobileCLIP2',
            'path': os.path.join(model_dir, 'mobileclip_models'),  # MobileCLIP needs subdirectory
            'feature_weight': 0.8,
            'description_weight': 0.2,
            'compositional_weight': 0.3,
            'adaptive_weight': 0.2,
            'image_weight': 0.6,
            'text_weight': 0.4
        }

    ]
    return all_models

def filter_models_by_selection(all_models, selected_model_types):
    filtered_models = []
    
    for model_type in selected_model_types:
        model_found = False
        for model in all_models:
            if model['type'] == model_type:
                filtered_models.append(model)
                model_found = True
                break
        
        if not model_found:
            print(f"Warning: Model '{model_type}' not found in available models")
    
    return filtered_models

def evaluate_clean_accuracy(model_type, model_path, dataset, args):

    # Create clean results directory (separate from attack results)
    # Underscore prefix ensures it appears at the top in directory listings
    clean_results_dir = os.path.join(args.output_dir, '_clean_results')
    os.makedirs(clean_results_dir, exist_ok=True)

    # Result file for clean accuracy
    clean_result_file = os.path.join(clean_results_dir, f"{model_type}_clean_{dataset}.txt")
    cache_file = os.path.join(clean_results_dir, f"{model_type}_{dataset}_clean.json")

    # Check if clean accuracy result already exists (unless --force_clean is set)
    if os.path.exists(cache_file) and not args.force_clean:
        try:
            with open(cache_file, 'r') as f:
                cache_data = json.load(f)
                print(f"Found cached clean accuracy for {model_type} on {dataset}: {cache_data['clean_accuracy']:.2f}%")
                print(f"  (using cached result from: {clean_result_file})")
                return cache_data['clean_accuracy']
        except (json.JSONDecodeError, KeyError) as e:
            print(f"Error reading cache file {cache_file}: {e}. Will re-evaluate.")

    if args.force_clean and os.path.exists(cache_file):
        print(f"--force_clean set: Ignoring cached result and re-evaluating...")

    print(f"Evaluating clean accuracy for {model_type} on {dataset}...")

    # Build command for clean accuracy evaluation
    cmd = [
        'python', 'run_attack.py',
        '--model_path', str(model_path) if model_path is not None else 'None',
        '--model_type', model_type,
        '--arch', args.arch,
        '--attack', 'clean',  # Special mode for clean accuracy only
        '--dataset', dataset,
        '--output_dir', clean_results_dir,  # Save to clean_results subfolder
        '--data_path', args.data_path,
        '--result_filename', f"{model_type}_clean_{dataset}.txt"  # Explicit filename
    ]
    
    # Execute the command
    print("Running command:", ' '.join(cmd))
    try:
        result = subprocess.run(cmd, check=True)
        print(f"Successfully evaluated clean accuracy for {model_type} on {dataset}")

        # Parse clean accuracy from the result file
        clean_accuracy = None
        if os.path.exists(clean_result_file):
            with open(clean_result_file, 'r') as f:
                for line in f:
                    if line.startswith("Clean accuracy:"):
                        clean_accuracy = float(line.split(":")[1].strip().replace("%", ""))
                        break

        if clean_accuracy is not None:
            # Cache the result for future use
            cache_data = {
                'model_type': model_type,
                'dataset': dataset,
                'clean_accuracy': clean_accuracy,
                'timestamp': time.time()
            }
            with open(cache_file, 'w') as f:
                json.dump(cache_data, f, indent=2)

            print(f"✓ Cached clean accuracy: {clean_accuracy:.2f}% → {cache_file}")
            return clean_accuracy
        else:
            print(f"⚠ Warning: Could not parse clean accuracy from {clean_result_file}")
            return None

    except subprocess.CalledProcessError as e:
        print(f"✗ Error evaluating clean accuracy for {model_type} on {dataset}: {e}")
        if e.stdout:
            print("STDOUT:", e.stdout)
        if e.stderr:
            print("STDERR:", e.stderr)
        return None


def get_result_filename(model_type, attack, dataset, epsilon_float):
    if attack == "clean":
        return f"{model_type}_clean_{dataset}.txt"
    return f"{model_type}_{dataset}_{attack}_eps{epsilon_float:.4f}.txt"


def check_result_exists(output_dir, model_type, dataset, attack, epsilon_float):
    result_file = os.path.join(output_dir, get_result_filename(model_type, attack, dataset, epsilon_float))
    return os.path.exists(result_file)


def main():
    args = parse_args()
    
    # Load configuration from YAML file
    config = load_config(args.config)
    
    # Apply command-line overrides
    if args.models:
        config['models'] = args.models
    if args.datasets:
        config['datasets'] = args.datasets
    if args.attacks:
        config['attacks'] = args.attacks
    if args.epsilons:
        config['epsilons'] = args.epsilons
    
    # Validate configuration
    if not validate_config(config):
        print("\nPlease fix configuration errors and try again.")
        return
    
    # Show config and exit if requested
    if args.show_config:
        print_config(config)
        return
    
    print_config(config)
    
    # Extract config values
    model_names = config['models']
    datasets = config['datasets']
    attacks = config.get('attacks', ['pgd', 'CW', 'sapa'])
    epsilons = config['epsilons']  # In 1/255 units
    output_dir = config['output']['output_dir']
    model_dir = config['model_settings']['model_dir']
    data_path = config['data']['data_path']
    num_test_samples = config['data'].get('num_test_samples', 100)  # Default to 100 samples
    arch = config['model_settings']['arch']
    visualize = config['output'].get('visualize', False)
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Get all available models and filter by user selection
    all_models = get_available_models(model_dir)
    models = filter_models_by_selection(all_models, model_names)
    
    print(f"\nSelected models for evaluation: {[m['type'] for m in models]}")
    print(f"Selected attacks: {attacks}")
    print(f"Selected datasets: {datasets}")
    print(f"Epsilon values: {[f'{e}/255' for e in epsilons]}")
    
    # Determine skip behavior
    skip_existing = args.skip_existing and not args.force
    
    # Step 1: Evaluate clean accuracy for all model-dataset combinations
    print("\n" + "="*80)
    print("STEP 1: Evaluating Clean Accuracy for All Model-Dataset Combinations")
    print("="*80)
    
    # Create a mock args object for evaluate_clean_accuracy
    class ConfigArgs:
        pass

    config_args = ConfigArgs()
    config_args.output_dir = output_dir
    config_args.data_path = data_path
    config_args.arch = arch
    config_args.force_clean = args.force_clean if hasattr(args, 'force_clean') else False
    
    clean_accuracies = {}
    for model in models:
        model_type = model['type']
        model_path = model['path']
        
        for dataset in datasets:
            print(f"\n{'-'*60}")
            print(f"Evaluating clean accuracy: {model_type} on {dataset}")
            print(f"{'-'*60}")
            
            clean_acc = evaluate_clean_accuracy(model_type, model_path, dataset, config_args)
            clean_accuracies[(model_type, dataset)] = clean_acc
    
    # Step 2: Evaluate adversarial attacks for all epsilon values
    print("\n" + "="*80)
    print("STEP 2: Evaluating Adversarial Attacks")
    print("="*80)
    
    # Track progress
    total_configs = len(models) * len(datasets) * len(attacks) * len(epsilons)
    completed = 0
    skipped = 0
    failed = []
    
    # Iterate over all combinations
    for model in models:
        model_type = model['type']
        model_path = model['path']
        
        for dataset in datasets:
            for epsilon in epsilons:
                epsilon_float = epsilon / 255.0
                
                print(f"\n\n{'='*80}")
                print(f"Evaluating: {model_type} on {dataset} with ε={epsilon}/255 ({epsilon_float:.6f})")
                print(f"Clean accuracy (cached): {clean_accuracies.get((model_type, dataset), 'N/A')}")
                print(f"{'='*80}")
                
                for attack in attacks:
                    completed += 1
                    
                    # Check if result already exists
                    if skip_existing and check_result_exists(output_dir, model_type, dataset, attack, epsilon_float):
                        print(f"\n⏭ Skipping {attack} attack (result exists): {model_type}/{dataset}/eps{epsilon_float:.4f}")
                        skipped += 1
                        continue
                    
                    print(f"\n{'-'*80}")
                    print(f"Running {attack} attack ({completed}/{total_configs})")
                    print(f"{'-'*80}")
                    
                    # No subdirectory needed, ensure output_dir exists
                    os.makedirs(output_dir, exist_ok=True)
                    
                    # Build command
                    attack_iters = get_attack_iterations(attack, config)

                    # Result filename includes epsilon
                    result_filename = get_result_filename(model_type, attack, dataset, epsilon_float)
                    cmd = [
                        'python', 'run_attack.py',
                        '--model_path', str(model_path) if model_path is not None else 'None',
                        '--model_type', model_type,
                        '--arch', arch,
                        '--attack', attack,
                        '--epsilon', str(epsilon_float),
                        '--dataset', dataset,
                        '--output_dir', output_dir,
                        '--data_path', data_path,
                        '--attack_iters', str(attack_iters),
                        '--skip_clean_eval',
                        '--evaluate_semantic_control',
                        '--result_filename', result_filename  # Custom result filename
                    ]
                    
                    # Add num_test_samples if specified
                    if num_test_samples is not None:
                        cmd.extend(['--num_test_samples', str(num_test_samples)])

                    # Add attack-specific parameters (with epsilon for auto-stepsize computation)
                    attack_params = get_attack_specific_params(attack, config, epsilon_float=epsilon_float)
                    cmd.extend(attack_params)
                    
                    # Add visualization flag if enabled
                    if visualize:
                        cmd.append('--visualize')
                    
                    # Add debug flag if enabled
                    debug_mode = config['output'].get('debug', False)
                    if debug_mode:
                        cmd.append('--debug')
                    
                    # Add cached clean accuracy
                    cached_clean_acc = clean_accuracies.get((model_type, dataset))
                    if cached_clean_acc is not None:
                        cmd.extend(['--cached_clean_accuracy', str(cached_clean_acc)])
                    
                    # Execute the command
                    print("Running command:", ' '.join(cmd))
                    try:
                        subprocess.run(cmd, check=True)
                        print(f"✓ Successfully completed {attack} attack")
                    except subprocess.CalledProcessError as e:
                        print(f"✗ Error running {attack}: {e}")
                        failed.append((model_type, dataset, attack, epsilon))
    
    # Print progress summary
    print("\n" + "="*80)
    print("ATTACK RUNS COMPLETE")
    print("="*80)
    print(f"Total configurations: {total_configs}")
    print(f"Completed: {completed - skipped - len(failed)}")
    print(f"Skipped (existing): {skipped}")
    print(f"Failed: {len(failed)}")
    
    if failed:
        print("\nFailed configurations:")
        for m, d, a, e in failed:
            print(f"  - {m} / {d} / {a} / ε={e}/255")
    
    # Step 3: Generate summary report
    print("\n" + "="*80)
    print("STEP 3: Generating Summary Report")
    print("="*80)
    
    generate_summary_report_yaml(models, datasets, attacks, epsilons, config, clean_accuracies)
    
    # Step 4: Export to Excel
    print("\n" + "="*80)
    print("STEP 4: Exporting Results to Excel")
    print("="*80)
    
    excel_file = export_summary_to_excel_yaml(models, datasets, attacks, epsilons, config, clean_accuracies)
    
    print(f"\n Comprehensive Excel analysis exported!")
    print(f" File location: {excel_file}")
    print(f" Open in Excel/LibreOffice for interactive analysis")
    
    print("\n✓ Evaluation complete!")


def generate_summary_report_yaml(models, datasets, attacks, epsilons, config, clean_accuracies):

    output_dir = config['output']['output_dir']
    summary_file = os.path.join(output_dir, '_summary_report.txt')
    
    # Collect all results data
    all_results = []
    
    for model in models:
        model_type = model['type']
        
        for dataset in datasets:
            clean_acc = clean_accuracies.get((model_type, dataset))
            
            for epsilon in epsilons:
                epsilon_float = epsilon / 255.0
                
                for attack in attacks:
                    result_filename = get_result_filename(model_type, attack, dataset, epsilon_float)
                    results_file = os.path.join(output_dir, result_filename)
                    
                    adv_acc = ssp = src = interpretability_score = sta = sta_alignment = None
                    
                    if os.path.exists(results_file):
                        try:
                            with open(results_file, 'r') as rf:
                                for line in rf:
                                    line = line.strip()
                                    if line.startswith("Adversarial accuracy:"):
                                        adv_acc = float(line.split(":")[1].strip().replace("%", ""))
                                    elif line.startswith("Semantic Similarity Preservation (SSP):"):
                                        ssp = float(line.split(":")[1].strip())
                                    elif line.startswith("Semantic Rank Correlation (SRC):"):
                                        src = float(line.split(":")[1].strip())
                                    elif line.startswith("Interpretability Score (IS):"):
                                        interpretability_score = float(line.split(":")[1].strip())
                                    elif line.startswith("Semantic Targeting Accuracy (STAccuracy):"):
                                        sta_str = line.split(":")[1].strip()
                                        if sta_str not in ["N/A (no semantic targeting)", "N/A"]:
                                            sta = float(sta_str)
                                    elif line.startswith("Semantic Target Alignment (STA):"):
                                        sta_align_str = line.split(":")[1].strip()
                                        if sta_align_str not in ["N/A", ""]:
                                            sta_alignment = float(sta_align_str)
                        except Exception as e:
                            print(f"Error reading {results_file}: {e}")
                    
                    robustness_drop = clean_acc - adv_acc if (clean_acc is not None and adv_acc is not None) else None
                    
                    all_results.append({
                        'model': model_type,
                        'epsilon': epsilon,
                        'epsilon_float': epsilon_float,
                        'dataset': dataset,
                        'attack': attack,
                        'clean_accuracy': clean_acc,
                        'adversarial_accuracy': adv_acc,
                        'robustness_drop': robustness_drop,
                        'ssp': ssp,
                        'src': src,
                        'interpretability_score': interpretability_score,
                        'sta': sta,
                        'sta_alignment': sta_alignment
                    })
    
    # Write summary report
    with open(summary_file, 'w') as f:
        f.write("ADVERSARIAL ROBUSTNESS EVALUATION SUMMARY\n")
        f.write("="*100 + "\n\n")
        f.write(f"Configuration:\n")
        f.write(f"- Architecture: {config['model_settings']['arch']}\n")
        f.write(f"- Epsilon values: {[f'{e}/255' for e in epsilons]}\n")
        f.write(f"- Attacks: {', '.join(attacks)}\n")
        f.write(f"- Datasets: {', '.join(datasets)}\n\n")
        
        # Detailed results table
        f.write("DETAILED RESULTS TABLE\n")
        f.write("="*200 + "\n")
        
        header = f"{'Model':<10} | {'ε (1/255)':<10} | {'Dataset':<12} | {'Attack':<12} | {'Clean %':<8} | {'Adv %':<8} | {'Drop %':<8} | {'SSP':<6} | {'SRC':<6} | {'IS':<6} | {'STAcc':<6} | {'STA':<6}"
        f.write(header + "\n")
        f.write("-" * len(header) + "\n")
        
        for r in all_results:
            clean_str = f"{r['clean_accuracy']:.2f}" if r['clean_accuracy'] else "N/A"
            adv_str = f"{r['adversarial_accuracy']:.2f}" if r['adversarial_accuracy'] else "N/A"
            drop_str = f"{r['robustness_drop']:.2f}" if r['robustness_drop'] else "N/A"
            ssp_str = f"{r['ssp']:.3f}" if r['ssp'] else "N/A"
            src_str = f"{r['src']:.3f}" if r['src'] else "N/A"
            is_str = f"{r['interpretability_score']:.3f}" if r['interpretability_score'] else "N/A"
            sta_str = f"{r['sta']:.3f}" if r['sta'] else "N/A"
            sta_align_str = f"{r['sta_alignment']:.3f}" if r['sta_alignment'] else "N/A"
            
            row = f"{r['model']:<10} | {r['epsilon']:<10} | {r['dataset']:<12} | {r['attack']:<12} | {clean_str:<8} | {adv_str:<8} | {drop_str:<8} | {ssp_str:<6} | {src_str:<6} | {is_str:<6} | {sta_str:<6} | {sta_align_str:<6}"
            f.write(row + "\n")
    
    # Save JSON data
    json_file = os.path.join(output_dir, '_summary_data.json')
    with open(json_file, 'w') as f:
        json.dump({
            'config': {
                'arch': config['model_settings']['arch'],
                'epsilons': epsilons,
                'attacks': attacks,
                'datasets': datasets
            },
            'results': all_results
        }, f, indent=2)
    
    print(f"Summary report saved to: {summary_file}")
    print(f"JSON data saved to: {json_file}")
    
    return all_results


def export_summary_to_excel_yaml(models, datasets, attacks, epsilons, config, clean_accuracies):
    
    output_dir = config['output']['output_dir']
    os.makedirs(output_dir, exist_ok=True)
    
    # Collect all results
    all_results = []
    
    for model in models:
        model_type = model['type']
        
        for dataset in datasets:
            clean_acc = clean_accuracies.get((model_type, dataset))
            
            for epsilon in epsilons:
                epsilon_float = epsilon / 255.0
                
                for attack in attacks:
                    result_filename = get_result_filename(model_type, attack, dataset, epsilon_float)
                    results_file = os.path.join(output_dir, result_filename)
                    
                    adv_acc = ssp = src = interpretability_score = sta = sta_alignment = None
                    
                    if os.path.exists(results_file):
                        try:
                            with open(results_file, 'r') as rf:
                                for line in rf:
                                    line = line.strip()
                                    if line.startswith("Adversarial accuracy:"):
                                        adv_acc = float(line.split(":")[1].strip().replace("%", ""))
                                    elif line.startswith("Semantic Similarity Preservation (SSP):"):
                                        ssp = float(line.split(":")[1].strip())
                                    elif line.startswith("Semantic Rank Correlation (SRC):"):
                                        src = float(line.split(":")[1].strip())
                                    elif line.startswith("Interpretability Score (IS):"):
                                        interpretability_score = float(line.split(":")[1].strip())
                                    elif line.startswith("Semantic Targeting Accuracy (STAccuracy):"):
                                        sta_str = line.split(":")[1].strip()
                                        if sta_str not in ["N/A (no semantic targeting)", "N/A"]:
                                            sta = float(sta_str)
                                    elif line.startswith("Semantic Target Alignment (STA):"):
                                        sta_align_str = line.split(":")[1].strip()
                                        if sta_align_str not in ["N/A", ""]:
                                            sta_alignment = float(sta_align_str)
                        except Exception as e:
                            print(f"Error reading {results_file}: {e}")
                    
                    robustness_drop = clean_acc - adv_acc if (clean_acc is not None and adv_acc is not None) else None
                    attack_success = (100 - adv_acc) if adv_acc is not None else None
                    
                    all_results.append({
                        'Model': model_type,
                        'Dataset': dataset,
                        'Attack': attack,
                        'Epsilon': epsilon,
                        'Epsilon_Float': epsilon_float,
                        'Clean_Accuracy': clean_acc,
                        'Adversarial_Accuracy': adv_acc,
                        'Robustness_Drop': robustness_drop,
                        'Attack_Success_Rate': attack_success,
                        'SSP': ssp,
                        'SRC': src,
                        'IS': interpretability_score,
                        'STAccuracy': sta,
                        'STA': sta_alignment
                    })
    
    df_main = pd.DataFrame(all_results)
    
    # Excel filename includes epsilon range
    eps_str = f"{min(epsilons)}-{max(epsilons)}" if len(epsilons) > 1 else str(epsilons[0])
    excel_file = os.path.join(output_dir, f'_summary_report_eps{eps_str}.xlsx')
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Sheet 1: All Results
        df_main.to_excel(writer, sheet_name='All_Results', index=False)
        
        # Sheet 2: Results by Epsilon (pivot table)
        if not df_main.empty and len(epsilons) > 1:
            for metric in ['Adversarial_Accuracy', 'Robustness_Drop', 'SSP', 'STA']:
                if metric in df_main.columns and df_main[metric].notna().any():
                    pivot = df_main.pivot_table(
                        values=metric,
                        index=['Model', 'Dataset', 'Attack'],
                        columns='Epsilon',
                        aggfunc='mean'
                    )
                    pivot.to_excel(writer, sheet_name=f'{metric[:15]}_by_Eps')
        
        # Sheet 3: Model Comparison
        model_stats = []
        for model in models:
            model_type = model['type']
            model_results = df_main[df_main['Model'] == model_type]
            
            if not model_results.empty:
                stats = {
                    'Model': model_type,
                    'Experiments': len(model_results),
                    'Avg_Clean_Acc': model_results['Clean_Accuracy'].mean(),
                    'Avg_Adv_Acc': model_results['Adversarial_Accuracy'].mean(),
                    'Avg_Robustness_Drop': model_results['Robustness_Drop'].mean(),
                    'Avg_SSP': model_results['SSP'].mean(),
                    'Avg_SRC': model_results['SRC'].mean(),
                    'Avg_IS': model_results['IS'].mean(),
                    'Avg_STAccuracy': model_results['STAccuracy'].mean(),
                    'Avg_STA': model_results['STA'].mean(),
                }
                model_stats.append(stats)
        
        if model_stats:
            pd.DataFrame(model_stats).to_excel(writer, sheet_name='Model_Comparison', index=False)
        
        # Sheet 4: Attack Comparison
        attack_stats = []
        for attack in attacks:
            attack_results = df_main[df_main['Attack'] == attack]
            
            if not attack_results.empty:
                stats = {
                    'Attack': attack,
                    'Experiments': len(attack_results),
                    'Avg_Attack_Success': attack_results['Attack_Success_Rate'].mean(),
                    'Avg_Robustness_Drop': attack_results['Robustness_Drop'].mean(),
                    'Avg_SSP': attack_results['SSP'].mean(),
                    'Avg_SRC': attack_results['SRC'].mean(),
                    'Avg_IS': attack_results['IS'].mean(),
                    'Avg_STAccuracy': attack_results['STAccuracy'].mean(),
                    'Avg_STA': attack_results['STA'].mean(),
                }
                attack_stats.append(stats)
        
        if attack_stats:
            pd.DataFrame(attack_stats).to_excel(writer, sheet_name='Attack_Comparison', index=False)
        
        # Sheet 5: Epsilon Impact Analysis
        if len(epsilons) > 1:
            eps_stats = []
            for eps in epsilons:
                eps_results = df_main[df_main['Epsilon'] == eps]
                
                if not eps_results.empty:
                    stats = {
                        'Epsilon': eps,
                        'Epsilon_Float': eps / 255.0,
                        'Experiments': len(eps_results),
                        'Avg_Adv_Acc': eps_results['Adversarial_Accuracy'].mean(),
                        'Avg_Robustness_Drop': eps_results['Robustness_Drop'].mean(),
                        'Avg_Attack_Success': eps_results['Attack_Success_Rate'].mean(),
                        'Avg_SSP': eps_results['SSP'].mean(),
                        'Avg_STA': eps_results['STA'].mean(),
                    }
                    eps_stats.append(stats)
            
            if eps_stats:
                pd.DataFrame(eps_stats).to_excel(writer, sheet_name='Epsilon_Impact', index=False)
    
    # Apply formatting
    try:
        format_excel_file(excel_file)
    except Exception as e:
        print(f"Warning: Could not apply Excel formatting: {e}")
    
    print(f"Excel report saved to: {excel_file}")
    return excel_file


# ============================================================================
# Legacy functions (kept for backward compatibility)
# ============================================================================

def generate_summary_report(models, datasets, attacks, args, clean_accuracies):

    summary_file = os.path.join(args.output_dir, '_summary_report.txt')
    summary_data = {}
    
    # Collect all results data first
    all_results = []
    
    for model in models:
        model_type = model['type']
        
        for dataset in datasets:
            # Get clean accuracy
            clean_acc = clean_accuracies.get((model_type, dataset))
            
            for attack in attacks:
                output_subdir = os.path.join(args.output_dir, f"{model_type}_{dataset}")
                results_file = os.path.join(output_subdir, f"{model_type}_{attack}_results.txt")
                
                adv_acc = None
                adv_acc = None
                ssp = None
                src = None
                interpretability_score = None
                sta = None

                if os.path.exists(results_file):
                    try:
                        with open(results_file, 'r') as rf:
                            for result_line in rf:
                                if result_line.startswith("Adversarial accuracy:"):
                                    adv_acc = float(result_line.split(":")[1].strip().replace("%", ""))
                                elif result_line.startswith("Semantic Similarity Preservation (SSP):"):
                                    ssp = float(result_line.split(":")[1].strip())
                                elif result_line.startswith("Semantic Rank Correlation (SRC):"):
                                    src = float(result_line.split(":")[1].strip())
                                elif result_line.startswith("Interpretability Score (IS):"):
                                    interpretability_score = float(result_line.split(":")[1].strip())
                                elif result_line.startswith("Semantic Targeting Accuracy (STA):"):
                                    sta_str = result_line.split(":")[1].strip()
                                    if sta_str not in ["N/A (no semantic targeting)", "N/A"]:
                                        sta = float(sta_str)
                    except Exception as e:
                        print(f"Error reading {results_file}: {e}")
                
                # Calculate robustness drop
                robustness_drop = clean_acc - adv_acc if (clean_acc is not None and adv_acc is not None) else None
                
                # Store result
                result_entry = {
                    'model': model_type,
                    'epsilon': args.epsilon,
                    'dataset': dataset,
                    'attack': attack,
                    'clean_accuracy': clean_acc,
                    'adversarial_accuracy': adv_acc,
                    'robustness_drop': robustness_drop,
                    'ssp': ssp,
                    'src': src,
                    'interpretability_score': interpretability_score,
                    'sta': sta
                }
                all_results.append(result_entry)
    
    # Write summary report
    with open(summary_file, 'w') as f:
        f.write("ADVERSARIAL ROBUSTNESS EVALUATION SUMMARY\n")
        f.write("="*80 + "\n\n")
        f.write(f"Evaluation Parameters:\n")
        f.write(f"- Architecture: {args.arch}\n")
        f.write(f"- Epsilon: {args.epsilon}\n")
        f.write(f"- Data Path: {args.data_path}\n")
        f.write(f"- Attacks Evaluated: {', '.join(attacks)}\n")
        f.write(f"- Datasets Evaluated: {', '.join(datasets)}\n\n")
        
        # Generate enhanced summary table
        f.write("DETAILED RESULTS TABLE\n")
        f.write("="*180 + "\n")
        
        # Table header
        header = f"{'Model':<10} | {'Epsilon':<8} | {'Dataset':<12} | {'Attack':<12} | {'Clean Acc (%)':<12} | {'Adv Acc (%)':<12} | {'Rob Drop (%)':<12} | {'SSP':<6} | {'SRC':<6} | {'IS':<6} | {'STA':<6}"
        f.write(header + "\n")
        f.write("-" * len(header) + "\n")
        
        # Table rows
        for result in all_results:
            model = result['model']
            epsilon = result['epsilon']
            dataset = result['dataset']
            attack = result['attack']
            clean_acc = result['clean_accuracy']
            adv_acc = result['adversarial_accuracy']
            rob_drop = result['robustness_drop']
            ssp = result['ssp']
            src = result['src']
            interpretability_score = result['interpretability_score']
            sta = result['sta']
            
            # Format values with 3 decimal places
            clean_str = f"{clean_acc:.3f}" if clean_acc is not None else "N/A"
            adv_str = f"{adv_acc:.3f}" if adv_acc is not None else "N/A"
            drop_str = f"{rob_drop:.3f}" if rob_drop is not None else "N/A"
            ssp_str = f"{ssp:.3f}" if ssp is not None else "N/A"
            src_str = f"{src:.3f}" if src is not None else "N/A"
            is_str = f"{interpretability_score:.3f}" if interpretability_score is not None else "N/A"
            sta_str = f"{sta:.3f}" if sta is not None else "N/A"

            row = (f"{model:<10} | {epsilon:<8} | {dataset:<12} | {attack:<12} | {clean_str:<9} | "
                  f"{adv_str:<8} | {drop_str:<8} | {ssp_str:<6} | {src_str:<6} | {is_str:<6} | {sta_str:<6}")
            f.write(row + "\n")
        
        f.write("\n")
        
        # Semantic metrics analysis by attack type
        f.write("SEMANTIC METRICS ANALYSIS BY ATTACK TYPE\n")
        f.write("="*80 + "\n")
        
        attack_semantic_stats = {}
        for attack in attacks:
            attack_results = [r for r in all_results if r['attack'] == attack]
            
            # Calculate averages for semantic metrics
            ssp_values = [r['ssp'] for r in attack_results if r['ssp'] is not None]
            src_values = [r['src'] for r in attack_results if r['src'] is not None]
            is_values = [r['interpretability_score'] for r in attack_results if r['interpretability_score'] is not None]
            sta_values = [r['sta'] for r in attack_results if r['sta'] is not None]
            
            attack_semantic_stats[attack] = {
                'avg_ssp': sum(ssp_values) / len(ssp_values) if ssp_values else None,
                'avg_src': sum(src_values) / len(src_values) if src_values else None,
                'avg_is': sum(is_values) / len(is_values) if is_values else None,
                'avg_sta': sum(sta_values) / len(sta_values) if sta_values else None,
                'num_samples': len(attack_results)
            }
        
        # Write semantic analysis table
        sem_header = f"{'Attack':<12} | {'Avg SSP':<8} | {'Avg SRC':<8} | {'Avg IS':<8} | {'Avg STA':<8} | {'Samples':<7}"
        f.write(sem_header + "\n")
        f.write("-" * len(sem_header) + "\n")
        
        for attack, stats in attack_semantic_stats.items():
            ssp_str = f"{stats['avg_ssp']:.3f}" if stats['avg_ssp'] is not None else "N/A"
            src_str = f"{stats['avg_src']:.3f}" if stats['avg_src'] is not None else "N/A"
            is_str = f"{stats['avg_is']:.3f}" if stats['avg_is'] is not None else "N/A"
            sta_str = f"{stats['avg_sta']:.3f}" if stats['avg_sta'] is not None else "N/A"
            
            sem_row = f"{attack:<12} | {ssp_str:<8} | {src_str:<8} | {is_str:<8} | {sta_str:<8} | {stats['num_samples']:<7}"
            f.write(sem_row + "\n")
        
        f.write("\n")

        # Generate aggregated statistics per model
        f.write("AGGREGATED STATISTICS PER MODEL\n")
        f.write("="*80 + "\n")
        
        model_stats = {}
        for model in models:
            model_type = model['type']
            model_results = [r for r in all_results if r['model'] == model_type]
            
            # Calculate average metrics
            clean_accs = [r['clean_accuracy'] for r in model_results if r['clean_accuracy'] is not None]
            adv_accs = [r['adversarial_accuracy'] for r in model_results if r['adversarial_accuracy'] is not None]
            rob_drops = [r['robustness_drop'] for r in model_results if r['robustness_drop'] is not None]
            ssp_values = [r['ssp'] for r in model_results if r['ssp'] is not None]
            src_values = [r['src'] for r in model_results if r['src'] is not None]
            is_values = [r['interpretability_score'] for r in model_results if r['interpretability_score'] is not None]
            sta_values = [r['sta'] for r in model_results if r['sta'] is not None]
            avg_clean = sum(clean_accs) / len(clean_accs) if clean_accs else None
            avg_adv = sum(adv_accs) / len(adv_accs) if adv_accs else None
            avg_drop = sum(rob_drops) / len(rob_drops) if rob_drops else None
            avg_ssp = sum(ssp_values) / len(ssp_values) if ssp_values else None
            avg_src = sum(src_values) / len(src_values) if src_values else None
            avg_is = sum(is_values) / len(is_values) if is_values else None
            avg_sta = sum(sta_values) / len(sta_values) if sta_values else None
            
            model_stats[model_type] = {
                'avg_clean': avg_clean,
                'avg_adv': avg_adv,
                'avg_drop': avg_drop,
                'avg_ssp': avg_ssp,
                'avg_src': avg_src,
                'avg_is': avg_is,
                'avg_sta': avg_sta,
                'num_experiments': len(model_results)
            }
        
        # Write aggregated statistics
        agg_header = f"{'Model':<10} | {'Avg Clean (%)':<13} | {'Avg Adv (%)':<13} | {'Avg Drop (%)':<13} | {'Avg SSP':<8} | {'Avg SRC':<8} | {'Avg IS':<8} | {'Avg STA':<8} | {'# Experiments':<13}"
        f.write(agg_header + "\n")
        f.write("-" * len(agg_header) + "\n")
        
        for model_type, stats in model_stats.items():
            avg_clean_str = f"{stats['avg_clean']:.3f}" if stats['avg_clean'] is not None else "N/A"
            avg_adv_str = f"{stats['avg_adv']:.3f}" if stats['avg_adv'] is not None else "N/A"
            avg_drop_str = f"{stats['avg_drop']:.3f}" if stats['avg_drop'] is not None else "N/A"
            avg_ssp_str = f"{stats['avg_ssp']:.3f}" if stats['avg_ssp'] is not None else "N/A"
            avg_src_str = f"{stats['avg_src']:.3f}" if stats['avg_src'] is not None else "N/A"
            avg_is_str = f"{stats['avg_is']:.3f}" if stats['avg_is'] is not None else "N/A"
            avg_sta_str = f"{stats['avg_sta']:.3f}" if stats['avg_sta'] is not None else "N/A"

            agg_row = f"{model_type:<10} | {avg_clean_str:<13} | {avg_adv_str:<13} | {avg_drop_str:<13} | {avg_ssp_str:<8} | {avg_src_str:<8} | {avg_is_str:<8} | {avg_sta_str:<8} | {stats['num_experiments']:<13}"
            f.write(agg_row + "\n")
        
        f.write("\n")
        
        # Generate per-dataset summary
        f.write("AGGREGATED STATISTICS PER DATASET\n")
        f.write("="*80 + "\n")
        
        dataset_stats = {}
        for dataset in datasets:
            dataset_results = [r for r in all_results if r['dataset'] == dataset]
            
            # Calculate average metrics
            clean_accs = [r['clean_accuracy'] for r in dataset_results if r['clean_accuracy'] is not None]
            adv_accs = [r['adversarial_accuracy'] for r in dataset_results if r['adversarial_accuracy'] is not None]
            rob_drops = [r['robustness_drop'] for r in dataset_results if r['robustness_drop'] is not None]
            ssp_values = [r['ssp'] for r in dataset_results if r['ssp'] is not None]
            src_values = [r['src'] for r in dataset_results if r['src'] is not None]
            is_values = [r['interpretability_score'] for r in dataset_results if r['interpretability_score'] is not None]
            sta_values = [r['sta'] for r in dataset_results if r['sta'] is not None]

            avg_clean = sum(clean_accs) / len(clean_accs) if clean_accs else None
            avg_adv = sum(adv_accs) / len(adv_accs) if adv_accs else None
            avg_drop = sum(rob_drops) / len(rob_drops) if rob_drops else None
            avg_ssp = sum(ssp_values) / len(ssp_values) if ssp_values else None
            avg_src = sum(src_values) / len(src_values) if src_values else None
            avg_is = sum(is_values) / len(is_values) if is_values else None
            avg_sta = sum(sta_values) / len(sta_values) if sta_values else None

            dataset_stats[dataset] = {
                'avg_clean': avg_clean,
                'avg_adv': avg_adv,
                'avg_drop': avg_drop,
                'avg_ssp': avg_ssp,
                'avg_src': avg_src,
                'avg_is': avg_is,
                'avg_sta': avg_sta,
                'num_experiments': len(dataset_results)
            }
        
        # Write dataset statistics
        dataset_header = f"{'Dataset':<12} | {'Avg Clean (%)':<13} | {'Avg Adv (%)':<13} | {'Avg Drop (%)':<13} | {'Avg SSP':<8} | {'Avg SRC':<8} | {'Avg IS':<8} | {'Avg STA':<8} | {'# Experiments':<13}"
        f.write(dataset_header + "\n")
        f.write("-" * len(dataset_header) + "\n")
        
        for dataset_name, stats in dataset_stats.items():
            avg_clean_str = f"{stats['avg_clean']:.3f}" if stats['avg_clean'] is not None else "N/A"
            avg_adv_str = f"{stats['avg_adv']:.3f}" if stats['avg_adv'] is not None else "N/A"
            avg_drop_str = f"{stats['avg_drop']:.3f}" if stats['avg_drop'] is not None else "N/A"
            avg_ssp_str = f"{stats['avg_ssp']:.3f}" if stats['avg_ssp'] is not None else "N/A"
            avg_src_str = f"{stats['avg_src']:.3f}" if stats['avg_src'] is not None else "N/A"
            avg_is_str = f"{stats['avg_is']:.3f}" if stats['avg_is'] is not None else "N/A"
            avg_sta_str = f"{stats['avg_sta']:.3f}" if stats['avg_sta'] is not None else "N/A"

            dataset_row = f"{dataset_name:<12} | {avg_clean_str:<13} | {avg_adv_str:<13} | {avg_drop_str:<13} | {avg_ssp_str:<8} | {avg_src_str:<8} | {avg_is_str:<8} | {avg_sta_str:<8} | {stats['num_experiments']:<13}"
            f.write(dataset_row + "\n")
    
    # Save summary data as JSON for further analysis
    json_summary_file = os.path.join(args.output_dir, '_summary_data.json')
    with open(json_summary_file, 'w') as f:
        json.dump({
            'parameters': {
                'architecture': args.arch,
                'epsilon': args.epsilon,
                'attacks': attacks,
                'datasets': datasets
            },
            'results': all_results,
            'model_statistics': model_stats,
            'dataset_statistics': dataset_stats
        }, f, indent=2)
    
    # Print summary to console
    print(f"\nSummary report saved to: {summary_file}")
    print(f"Summary data (JSON) saved to: {json_summary_file}")
    
    # Print the enhanced table to console as well
    print("\n" + "="*150)
    print("FINAL RESULTS TABLE")
    print("="*150)

    print(f"{'Model':<10} | {'Epsilon':<8} | {'Dataset':<12} | {'Attack':<12} | {'Clean Acc (%)':<12} | {'Adv Acc (%)':<12} | {'Rob Drop (%)':<12} | {'SSP':<6} | {'SRC':<6} | {'IS':<6} | {'STA':<6}")
    print("-" * 150)
    
    for result in all_results:
        model = result['model']
        epsilon = result['epsilon']
        dataset = result['dataset']
        attack = result['attack']
        clean_acc = result['clean_accuracy']
        adv_acc = result['adversarial_accuracy']
        rob_drop = result['robustness_drop']
        ssp = result['ssp']
        src = result['src']
        is_score = result['interpretability_score']
        sta = result['sta']

        # Format values with 3 decimal places
        clean_str = f"{clean_acc:.3f}" if clean_acc is not None else "N/A"
        adv_str = f"{adv_acc:.3f}" if adv_acc is not None else "N/A"
        drop_str = f"{rob_drop:.3f}" if rob_drop is not None else "N/A"
        ssp_str = f"{ssp:.3f}" if ssp is not None else "N/A"
        src_str = f"{src:.3f}" if src is not None else "N/A"
        is_str = f"{is_score:.3f}" if is_score is not None else "N/A"
        sta_str = f"{sta:.3f}" if sta is not None else "N/A"

        print(f"{model:<10} | {epsilon:<8.3f} | {dataset:<12} | {attack:<12} | {clean_str:<12} | {adv_str:<12} | {drop_str:<12} | {ssp_str:<6} | {src_str:<6} | {is_str:<6} | {sta_str:<6}")

def export_summary_to_excel(models, datasets, attacks, args, clean_accuracies, output_dir=None):
    if output_dir is None:
        output_dir = args.output_dir
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Collect all results data with semantic metrics
    all_results = []
    
    for model in models:
        model_type = model['type']
        
        for dataset in datasets:
            # Get clean accuracy
            clean_acc = clean_accuracies.get((model_type, dataset))
            
            for attack in attacks:
                output_subdir = os.path.join(output_dir, f"{model_type}_{dataset}")
                results_file = os.path.join(output_subdir, f"{model_type}_{attack}_results.txt")
                
                # Initialize all metrics
                adv_acc = ssp = src = interpretability_score = sta = None
                
                if os.path.exists(results_file):
                    try:
                        with open(results_file, 'r') as rf:
                            for result_line in rf:
                                line = result_line.strip()
                                if line.startswith("Adversarial accuracy:"):
                                    adv_acc = float(line.split(":")[1].strip().replace("%", ""))
                                elif line.startswith("Semantic Similarity Preservation (SSP):"):
                                    ssp = float(line.split(":")[1].strip())
                                elif line.startswith("Semantic Rank Correlation (SRC):"):
                                    src = float(line.split(":")[1].strip())
                                elif line.startswith("Interpretability Score (IS):"):
                                    interpretability_score = float(line.split(":")[1].strip())
                                elif line.startswith("Semantic Targeting Accuracy (STA):"):
                                    sta_str = line.split(":")[1].strip()
                                    if sta_str not in ["N/A (no semantic targeting)", "N/A"]:
                                        sta = float(sta_str)
                    except Exception as e:
                        print(f"Error reading {results_file}: {e}")
                
                # Calculate robustness drop and attack success rate
                robustness_drop = clean_acc - adv_acc if (clean_acc is not None and adv_acc is not None) else None
                attack_success_rate = (100 - adv_acc) if adv_acc is not None else None

                # Store result
                result_entry = {
                    'Model': model_type,
                    'Dataset': dataset,
                    'Attack': attack,
                    'Epsilon': args.epsilon,
                    'Clean_Accuracy': clean_acc,
                    'Adversarial_Accuracy': adv_acc,
                    'Robustness_Drop': robustness_drop,
                    'Attack_Success_Rate': attack_success_rate,
                    'SSP': ssp,
                    'SRC': src,
                    'Interpretability_Score': interpretability_score,
                    'STA': sta
                }
                all_results.append(result_entry)
    
    # Create DataFrames for different analysis perspectives
    df_main = pd.DataFrame(all_results)
    
    # Excel file path
    excel_file = os.path.join(output_dir, f'_summary_report_eps_{args.epsilon}.xlsx')
    
    with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
        # Sheet 1: Complete Results Table
        df_main.to_excel(writer, sheet_name='All_Results', index=False)
        
        # Sheet 2: Model Comparison (Aggregated Statistics)
        model_stats = []
        for model in models:
            model_type = model['type']
            model_results = df_main[df_main['Model'] == model_type]
            
            if not model_results.empty:
                stats = {
                    'Model': model_type,
                    'Num_Experiments': len(model_results),
                    'Avg_Clean_Accuracy': model_results['Clean_Accuracy'].mean(),
                    'Avg_Adversarial_Accuracy': model_results['Adversarial_Accuracy'].mean(),
                    'Avg_Robustness_Drop': model_results['Robustness_Drop'].mean(),
                    'Avg_Attack_Success_Rate': model_results['Attack_Success_Rate'].mean(),
                    'Avg_SSP': model_results['SSP'].mean(),
                    'Avg_SRC': model_results['SRC'].mean(),
                    'Avg_Interpretability_Score': model_results['Interpretability_Score'].mean(),
                    'Avg_STA': model_results['STA'].mean(),
                    'Min_Robustness_Drop': model_results['Robustness_Drop'].min(),
                    'Max_Robustness_Drop': model_results['Robustness_Drop'].max(),
                    'Std_Robustness_Drop': model_results['Robustness_Drop'].std()
                }
                model_stats.append(stats)
        
        df_model_stats = pd.DataFrame(model_stats)
        df_model_stats.to_excel(writer, sheet_name='Model_Comparison', index=False)
        
        # Sheet 3: Attack Comparison (Aggregated by Attack Type)
        attack_stats = []
        for attack in attacks:
            attack_results = df_main[df_main['Attack'] == attack]
            
            if not attack_results.empty:
                stats = {
                    'Attack': attack,
                    'Num_Experiments': len(attack_results),
                    'Avg_Attack_Success_Rate': attack_results['Attack_Success_Rate'].mean(),
                    'Avg_Robustness_Drop': attack_results['Robustness_Drop'].mean(),
                    'Avg_SSP': attack_results['SSP'].mean(),
                    'Avg_SRC': attack_results['SRC'].mean(),
                    'Avg_Interpretability_Score': attack_results['Interpretability_Score'].mean(),
                    'Avg_STA': attack_results['STA'].mean(),
                    'Max_Attack_Success_Rate': attack_results['Attack_Success_Rate'].max(),
                    'Min_Attack_Success_Rate': attack_results['Attack_Success_Rate'].min(),
                    'Std_Attack_Success_Rate': attack_results['Attack_Success_Rate'].std()
                }
                attack_stats.append(stats)
        
        df_attack_stats = pd.DataFrame(attack_stats)
        df_attack_stats.to_excel(writer, sheet_name='Attack_Comparison', index=False)
        
        # Sheet 4: Dataset Analysis
        dataset_stats = []
        for dataset in datasets:
            dataset_results = df_main[df_main['Dataset'] == dataset]
            
            if not dataset_results.empty:
                stats = {
                    'Dataset': dataset,
                    'Num_Experiments': len(dataset_results),
                    'Avg_Clean_Accuracy': dataset_results['Clean_Accuracy'].mean(),
                    'Avg_Adversarial_Accuracy': dataset_results['Adversarial_Accuracy'].mean(),
                    'Avg_Robustness_Drop': dataset_results['Robustness_Drop'].mean(),
                    'Avg_SSP': dataset_results['SSP'].mean(),
                    'Avg_SRC': dataset_results['SRC'].mean(),
                    'Avg_Interpretability_Score': dataset_results['Interpretability_Score'].mean(),
                    'Most_Robust_Model': dataset_results.loc[dataset_results['Robustness_Drop'].idxmin(), 'Model'] if not dataset_results['Robustness_Drop'].isna().all() else 'N/A',
                    'Least_Robust_Model': dataset_results.loc[dataset_results['Robustness_Drop'].idxmax(), 'Model'] if not dataset_results['Robustness_Drop'].isna().all() else 'N/A'
                }
                dataset_stats.append(stats)
        
        df_dataset_stats = pd.DataFrame(dataset_stats)
        df_dataset_stats.to_excel(writer, sheet_name='Dataset_Analysis', index=False)
        
        # Sheet 5: Semantic Analysis Focus
        semantic_results = df_main.dropna(subset=['SSP', 'SRC', 'Interpretability_Score'])
        if not semantic_results.empty:
            semantic_analysis = semantic_results[['Model', 'Dataset', 'Attack', 'SSP', 'SRC', 'Interpretability_Score', 'STA']].copy()
            
            # Add semantic rankings
            semantic_analysis['SSP_Rank'] = semantic_analysis['SSP'].rank(ascending=False, method='dense')
            semantic_analysis['SRC_Rank'] = semantic_analysis['SRC'].rank(ascending=False, method='dense')
            semantic_analysis['IS_Rank'] = semantic_analysis['Interpretability_Score'].rank(ascending=False, method='dense')
            
            semantic_analysis.to_excel(writer, sheet_name='Semantic_Analysis', index=False)
        
        # Sheet 6: SAPA-Specific Analysis (if SAPA attack exists)
        if 'sapa' in attacks:
            sapa_results = df_main[df_main['Attack'] == 'sapa'].copy()
            if not sapa_results.empty:
                sapa_results = sapa_results.dropna(subset=['STA'])
                if not sapa_results.empty:
                    # Add additional SAPA analysis
                    sapa_results['Semantic_Control_Effectiveness'] = sapa_results['STA'] * 100  # Convert to percentage
                    sapa_results['SAPA_Quality_Score'] = (sapa_results['Attack_Success_Rate'] * 0.6 + 
                                                        sapa_results['STA'] * 100 * 0.4)  # Combined score
                    
                    sapa_analysis = sapa_results[['Model', 'Dataset', 'Attack_Success_Rate', 'STA', 
                                                'Semantic_Control_Effectiveness', 'SAPA_Quality_Score', 
                                                'SSP', 'SRC', 'Interpretability_Score']].copy()
                    sapa_analysis.to_excel(writer, sheet_name='SAPA_Analysis', index=False)
        
        # Sheet 7: Model-Attack Interaction Matrix
        # Create pivot tables for easy comparison
        if not df_main.empty:
            # Robustness Drop Matrix (Models vs Attacks)
            robustness_matrix = df_main.pivot_table(
                values='Robustness_Drop', 
                index='Model', 
                columns='Attack', 
                aggfunc='mean'
            )
            robustness_matrix.to_excel(writer, sheet_name='Robustness_Matrix')
            
            # SSP Matrix (Models vs Attacks)
            ssp_matrix = df_main.pivot_table(
                values='SSP', 
                index='Model', 
                columns='Attack', 
                aggfunc='mean'
            )
            ssp_matrix.to_excel(writer, sheet_name='SSP_Matrix')
        
        # Sheet 8: Top/Bottom Performers
        if not df_main.empty:
            performance_analysis = []
            
            # Most robust model-dataset-attack combinations
            top_robust = df_main.nsmallest(10, 'Robustness_Drop')[['Model', 'Dataset', 'Attack', 'Robustness_Drop', 'SSP', 'Interpretability_Score']]
            
            # Least robust combinations
            least_robust = df_main.nlargest(10, 'Robustness_Drop')[['Model', 'Dataset', 'Attack', 'Robustness_Drop', 'SSP', 'Interpretability_Score']]
            
            # Highest semantic coherence
            high_ssp = df_main.nlargest(10, 'SSP')[['Model', 'Dataset', 'Attack', 'SSP', 'SRC', 'Interpretability_Score', 'Robustness_Drop']]
            
            # Export these analyses
            top_robust.to_excel(writer, sheet_name='Most_Robust', index=False)
            least_robust.to_excel(writer, sheet_name='Least_Robust', index=False)
            high_ssp.to_excel(writer, sheet_name='Highest_SSP', index=False)
    
    # Apply formatting to make the Excel file more readable
    format_excel_file(excel_file)
    
    print(f"Excel analysis exported to: {excel_file}")
    print(f"Sheets included:")
    print("  - All_Results: Complete results table")
    print("  - Model_Comparison: Aggregated model statistics")  
    print("  - Attack_Comparison: Aggregated attack statistics")
    print("  - Dataset_Analysis: Dataset-specific analysis")
    print("  - Semantic_Analysis: Semantic metrics focus")
    if 'sapa' in attacks:
        print("  - SAPA_Analysis: SAPA-specific semantic targeting analysis")
    print("  - Robustness_Matrix: Model vs Attack robustness heatmap")
    print("  - SSP_Matrix: Model vs Attack semantic similarity heatmap")
    print("  - Most_Robust: Top 10 most robust combinations")
    print("  - Least_Robust: Top 10 least robust combinations")
    print("  - Highest_SSP: Top 10 highest semantic similarity combinations")
    
    return excel_file

def format_excel_file(excel_file):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = load_workbook(excel_file)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Apply formatting to each worksheet
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Format headers (first row)
        if ws.max_row > 0:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 20)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Apply number formatting to numeric columns
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    if cell.column in [5, 6, 7]:  # Accuracy columns
                        cell.number_format = '0.00'
                    else:  # Other numeric columns
                        cell.number_format = '0.000'
    
    wb.save(excel_file)


if __name__ == "__main__":
    print("running....")
    main()
